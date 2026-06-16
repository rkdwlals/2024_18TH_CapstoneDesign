from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import send_alarm as s

#이걸로 쓰는 거에용
wb = Workbook()
ws = wb.active

# 기본 검사 정보 여기서 넣었어요
ws['A1'] = '이름'
ws['B1'] = '나이(만)'
ws['C1'] = '나잇대 평균 점수'
ws['D1'] = '검사자 점수'

# 임시로 넣은 거에요(평균은 자료 참조함)
ws['A2'] = '나응애'
ws['B2'] = 5
ws['C2'] = 80  # 나잇대 평균 점수
ws['D2'] = 76  # 검사자의 점수

# 셀 크기를 조정했습니다
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15

# 그림이 들어갈 열의 크기를 넗혀줬습니다
ws.column_dimensions['E'].width = 25  # 검사자의 그림이 들어가는 열
ws.column_dimensions['F'].width = 25  # 나잇대 평균 그림이 들어가는 열

# 검사자 그림과 나잇대 평균 그림 구별용 텍스트입니다
ws['E1'] = '검사자의 그림'
ws['F1'] = '나잇대 평균 그림'

# 그림이 들어가는 셀의 높이를 바꾸었습니다
ws.row_dimensions[2].height = 120  # 두 번째 행의 높이를 키워 그림이 들어갈 공간 확보

# 이미지 크기를 조정하기 위해 그림을 불러옵니다
examiner_image = PILImage.open('user.jpg')
average_image = PILImage.open('average.jpg')

# 이미지 크기 조정 (너비 150px, 높이 150px로 조정했습니다)
examiner_image = examiner_image.resize((150, 150))
average_image = average_image.resize((150, 150))

# 수정된 이미지를 임시 저장
examiner_image.save('resized_user.jpg')
average_image.save('resized_average.jpg')

# 이미지 삽입
img_examiner = Image('resized_user.jpg')  # 조정된 검사자의 그림
img_average = Image('resized_average.jpg')  # 조정된 나잇대 평균 그림

# 그림 삽입 위치 입니다
ws.add_image(img_examiner, 'E2')  # 검사자의 그림을 E2 셀에 삽입
ws.add_image(img_average, 'F2')   # 나잇대 평균 그림을 F2 셀에 삽입

# 엑셀 파일 저장
wb.save("examiner_report.xlsx")
s.send_excel_link()