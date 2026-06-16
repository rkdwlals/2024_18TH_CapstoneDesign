#!/usr/bin/env python
# coding: utf-8

import json
import os
import datetime
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

# 1. Kakao API 유틸리티 함수들
def save_tokens(filename, tokens):
    with open(filename, "w") as fp:
        json.dump(tokens, fp)

def load_tokens(filename):
    with open(filename) as fp:
        tokens = json.load(fp)
    return tokens

def update_tokens(app_key, filename):
    tokens = load_tokens(filename)

    url = "https://kauth.kakao.com/oauth/token"
    data = {
        "grant_type": "refresh_token",
        "client_id": app_key,
        "refresh_token": tokens['refresh_token']
    }
    response = requests.post(url, data=data)

    if response.status_code != 200:
        print("error! because ", response.json())
        tokens = None
    else:
        print(response.json())
        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = filename + "." + now
        os.rename(filename, backup_filename)
        tokens['access_token'] = response.json()['access_token']
        save_tokens(filename, tokens)
    return tokens

def send_message(filename, template):
    tokens = load_tokens(filename)

    headers = {
        "Authorization": "Bearer " + tokens['access_token']
    }

    payload = {
        "template_object": json.dumps(template)
    }

    url = "https://kapi.kakao.com/v2/api/talk/memo/default/send"
    res = requests.post(url, data=payload, headers=headers)

    return res


# 2. 엑셀 파일 생성 함수
def create_excel_file():
    wb = Workbook()
    ws = wb.active

    # 기본 검사 정보
    ws['A1'] = '이름'
    ws['B1'] = '나이(만)'
    ws['C1'] = '나잇대 평균 점수'
    ws['D1'] = '검사자 점수'

    # 임시로 넣은 데이터
    ws['A2'] = '나응애'
    ws['B2'] = 5
    ws['C2'] = 80  # 나잇대 평균 점수
    ws['D2'] = 76  # 검사자의 점수

    # 셀 크기 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    # 그림이 들어갈 열의 크기 조정
    ws.column_dimensions['E'].width = 25  # 검사자의 그림
    ws.column_dimensions['F'].width = 25  # 나잇대 평균 그림

    # 검사자 및 나잇대 평균 그림 구별용 텍스트
    ws['E1'] = '검사자의 그림'
    ws['F1'] = '나잇대 평균 그림'

    # 그림이 들어가는 셀의 높이 조정
    ws.row_dimensions[2].height = 120  # 두 번째 행 높이 조정

    # 이미지 크기 조정
    examiner_image = PILImage.open('user.jpg')
    average_image = PILImage.open('average.jpg')

    # 이미지 크기 조정 (150x150px)
    examiner_image = examiner_image.resize((150, 150))
    average_image = average_image.resize((150, 150))

    # 임시 저장
    examiner_image.save('resized_user.jpg')
    average_image.save('resized_average.jpg')

    # 이미지 삽입
    img_examiner = Image('resized_user.jpg')
    img_average = Image('resized_average.jpg')

    ws.add_image(img_examiner, 'E2')
    ws.add_image(img_average, 'F2')

    # 엑셀 파일 저장
    excel_filename = "examiner_report.xlsx"
    wb.save(excel_filename)
    return excel_filename


# 3. 엑셀 파일 링크 전송 함수
def send_excel_link():
    KAKAO_TOKEN_FILENAME = "kakao_token.json"  # 카카오 토큰 파일 경로
    KAKAO_APP_KEY = "c9c545f59d696c28f7e44352c65f089f"  # 카카오 앱 키

    # 토큰 갱신
    tokens = update_tokens(KAKAO_APP_KEY, KAKAO_TOKEN_FILENAME)

    # 엑셀 파일을 생성하고 링크 제공
    excel_filename = create_excel_file()

    # 엑셀 파일 다운로드 링크 포함 템플릿
    template = {
        "object_type": "text",
        "text": "엑셀 파일을 다운로드하려면 버튼을 클릭하세요!",
        "link": {
            "web_url": f"https://your-file-hosting-service.com/{excel_filename}",  # 파일 다운로드 링크
            "mobile_web_url": f"https://your-file-hosting-service.com/{excel_filename}"
        },
        "button_title": "엑셀 파일 다운로드"
    }

    # 메시지 전송
    res = send_message(KAKAO_TOKEN_FILENAME, template)
    if res.json().get('result_code') == 0:
        print('텍스트 메시지를 성공적으로 보냈습니다.')
    else:
        print('텍스트 메시지를 보내지 못했습니다. 오류 메시지 : ', res.json())


# 4. 실행
send_excel_link()
